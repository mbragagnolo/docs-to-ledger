"""Ledger workbook: read, deduplicate, and write transaction data."""
from __future__ import annotations

import datetime
import os
import tempfile
from pathlib import Path
from typing import Any, Literal

from docs_to_ledger.errors import OutputWriteError
from docs_to_ledger.model import MatchStatus, Transaction

# ---------------------------------------------------------------------------
# Column definitions
# ---------------------------------------------------------------------------

# Monthly sheet columns (in order)
MONTHLY_COLUMNS = [
    "Date",
    "Description",
    "Debit",
    "Credit",
    "Account",
    "Category",
    "Transaction ID",
    "Occurrence",
    "Document",
    "Archived Path",
    "Match Status",
    "Confidence",
    "Imported At",
]

# Review sheet columns
REVIEW_COLUMNS = ["Date", "Description", "Reason"]

# ---------------------------------------------------------------------------
# Helper: build column index from header row
# ---------------------------------------------------------------------------


def _col_index(headers: list[str | None]) -> dict[str, int]:
    """Return {column_name: 0-based_index} for known headers."""
    return {h: i for i, h in enumerate(headers) if h is not None}


# ---------------------------------------------------------------------------
# xlsx backend
# ---------------------------------------------------------------------------


class _XlsxBackend:
    """openpyxl-backed workbook operations."""

    def __init__(self) -> None:
        import openpyxl

        self._openpyxl = openpyxl
        self._wb: Any = None  # openpyxl.Workbook

    # ------------------------------------------------------------------
    # Open / create
    # ------------------------------------------------------------------

    def open_or_create(self, path: Path) -> None:
        openpyxl = self._openpyxl
        if path.exists():
            self._wb = openpyxl.load_workbook(str(path))
        else:
            self._wb = openpyxl.Workbook()
            # Remove the default empty sheet that openpyxl creates
            default = self._wb.active
            if default is not None:
                del self._wb[default.title]
    # ------------------------------------------------------------------
    # Sheet access helpers
    # ------------------------------------------------------------------

    def _sheet(self, name: str) -> Any:
        """Return existing sheet, creating it (with headers) if absent.

        Monthly sheets are migrated to the current MONTHLY_COLUMNS layout on first access.
        """
        wb = self._wb
        assert wb is not None
        if name not in wb.sheetnames:
            ws = wb.create_sheet(name)
            if name == "Review":
                ws.append(REVIEW_COLUMNS)
            else:
                ws.append(MONTHLY_COLUMNS)
        elif name != "Review":
            ws = wb[name]
            header_row = next(ws.iter_rows(min_row=1, max_row=1), None)
            if header_row is not None:
                current_headers = [cell.value for cell in header_row]
                if current_headers != MONTHLY_COLUMNS:
                    self._migrate_to_new_format(ws, current_headers)
        return wb[name]
    def _migrate_to_new_format(
        self, ws: Any, old_headers: list[object | None]
    ) -> None:
        """Migrate a sheet from a legacy column layout to the current MONTHLY_COLUMNS."""
        old_idx = {str(h): i for i, h in enumerate(old_headers) if h is not None}

        data_snapshots: list[dict[str, object]] = []
        for row in ws.iter_rows(min_row=2, values_only=True):
            snap: dict[str, object] = {
                col: row[i] for col, i in old_idx.items() if i < len(row)
            }
            data_snapshots.append(snap)

        for col_i, col_name in enumerate(MONTHLY_COLUMNS, start=1):
            ws.cell(row=1, column=col_i).value = col_name
        for row_num, snap in enumerate(data_snapshots, start=2):
            for col_i, col_name in enumerate(MONTHLY_COLUMNS, start=1):
                val: object = snap.get(col_name, "")
                if col_name == "Occurrence" and (val is None or val == ""):
                    val = 0
                ws.cell(row=row_num, column=col_i).value = val
    def _get_or_none(self, name: str) -> Any:
        wb = self._wb
        assert wb is not None
        if name in wb.sheetnames:
            return wb[name]
        return None

    def sheet_names(self) -> list[str]:
        assert self._wb is not None
        return list(self._wb.sheetnames)
    # ------------------------------------------------------------------
    # Read existing keys
    # ------------------------------------------------------------------

    def read_keys(self) -> dict[tuple[str, int], tuple[str, int]]:
        """Return {(txn_id, occurrence): (sheet_name, row_number)} for all monthly sheets."""
        result: dict[tuple[str, int], tuple[str, int]] = {}
        assert self._wb is not None
        for sheet_name in self._wb.sheetnames:
            if sheet_name == "Review":
                continue
            ws = self._wb[sheet_name]
            row_iter = ws.iter_rows(min_row=1, max_row=1)
            header_row = next(row_iter, None)
            if header_row is None:
                continue
            header = [cell.value for cell in header_row]
            col_idx = _col_index(header)

            txnid_col = col_idx.get("Transaction ID")
            occ_col = col_idx.get("Occurrence")

            for row_num, row in enumerate(ws.iter_rows(min_row=2, values_only=True), start=2):
                if txnid_col is None:
                    # Old format: try known positions
                    # Old format header: Date, Description, Debit, Credit,
                    # Account, Transaction ID, Imported At
                    # Transaction ID is at index 5 (0-based)
                    txn_id_val = row[5] if len(row) > 5 else None
                    occ_val = 0  # Old format had no occurrence column; default to 0
                else:
                    txn_id_val = row[txnid_col] if txnid_col < len(row) else None
                    occ_val = row[occ_col] if occ_col is not None and occ_col < len(row) else 0

                if txn_id_val is not None:
                    txn_id_str = str(txn_id_val)
                    occ_int = int(occ_val) if occ_val is not None else 0
                    result[(txn_id_str, occ_int)] = (sheet_name, row_num)
        return result

    # ------------------------------------------------------------------
    # Upsert
    # ------------------------------------------------------------------

    def upsert_row(
        self,
        txn: Transaction,
        sheet_name: str,
        existing_index: dict[tuple[str, int], tuple[str, int]],
    ) -> int:
        """Insert row if key not present. Returns 1 if added, 0 if skipped."""
        key = (txn.txn_id, txn.occurrence)
        if key in existing_index:
            return 0

        ws = self._sheet(sheet_name)
        now_str = datetime.datetime.now().isoformat(timespec="seconds")
        row_data = [
            txn.date.isoformat() if txn.date else "",
            txn.description,
            str(txn.debit) if txn.debit is not None else "",
            str(txn.credit) if txn.credit is not None else "",
            txn.account,
            "",  # Category — empty in v1
            txn.txn_id,
            txn.occurrence,
            "",  # Document
            "",  # Archived Path
            "",  # Match Status
            "",  # Confidence
            now_str,
        ]
        ws.append(row_data)
        # Update the in-memory index with the new row number
        new_row_num = ws.max_row
        existing_index[key] = (sheet_name, new_row_num)
        return 1

    # ------------------------------------------------------------------
    # set_match
    # ------------------------------------------------------------------

    def set_match(
        self,
        key: tuple[str, int],
        document_link: str,
        archived_path: str,
        status: MatchStatus,
        confidence: float,
        existing_index: dict[tuple[str, int], tuple[str, int]],
    ) -> None:
        if key not in existing_index:
            return
        sheet_name, row_num = existing_index[key]
        ws = self._get_or_none(sheet_name)
        if ws is None:
            return
        row_iter = ws.iter_rows(min_row=1, max_row=1)
        header_row = next(row_iter, None)
        if header_row is None:
            return
        header = [cell.value for cell in header_row]
        col_idx = _col_index(header)

        def _set(col_name: str, value: object) -> None:
            c = col_idx.get(col_name)
            if c is not None:
                cell = ws.cell(row=row_num, column=c + 1)
                cell.value = value
        _set("Document", document_link)
        _set("Archived Path", archived_path)
        _set("Match Status", status.value)
        _set("Confidence", confidence)

    # ------------------------------------------------------------------
    # Review sheet
    # ------------------------------------------------------------------

    def add_to_review(self, items: list[str]) -> None:
        ws = self._sheet("Review")
        now_str = datetime.datetime.now().isoformat(timespec="seconds")
        for item in items:
            ws.append(["", "", item, now_str])
    # ------------------------------------------------------------------
    # Save
    # ------------------------------------------------------------------

    def save(self, target: Path) -> None:
        assert self._wb is not None
        self._wb.save(str(target))

# ---------------------------------------------------------------------------
# ods backend
# ---------------------------------------------------------------------------


class _OdsBackend:
    """odfpy-backed workbook operations (simplified; no hyperlinks in v1)."""

    def __init__(self) -> None:
        from odf import text as odf_text
        from odf.table import Table, TableCell, TableRow

        self._odf_text = odf_text
        self._Table = Table
        self._TableRow = TableRow
        self._TableCell = TableCell

        # In-memory store: {sheet_name: list[list[object]]}
        # Row 0 is always the header.
        self._sheets: dict[str, list[list[object]]] = {}

    # ------------------------------------------------------------------
    # Open / create
    # ------------------------------------------------------------------

    def open_or_create(self, path: Path) -> None:
        if path.exists():
            self._load(path)
        # If not found, _sheets starts empty (we'll create sheets on demand)

    @staticmethod
    def _element_text(element: Any) -> str:
        """Extract plain text from an odfpy element by walking its child nodes."""
        parts: list[str] = []
        for child in element.childNodes:
            if hasattr(child, "data"):
                parts.append(child.data)
            elif hasattr(child, "childNodes"):
                parts.append(_OdsBackend._element_text(child))
        return "".join(parts)

    def _load(self, path: Path) -> None:
        from odf.opendocument import load as odf_load

        doc = odf_load(str(path))
        sheets = doc.spreadsheet.getElementsByType(self._Table)
        for sheet in sheets:
            name = sheet.getAttribute("name")
            rows_data: list[list[object]] = []
            for row in sheet.getElementsByType(self._TableRow):
                cells = row.getElementsByType(self._TableCell)
                row_vals: list[object] = []
                for cell in cells:
                    # Get text content from child <text:p> elements
                    p_elements = cell.getElementsByType(self._odf_text.P)
                    val: object = ""
                    if p_elements:
                        val = self._element_text(p_elements[0])
                    row_vals.append(val)
                rows_data.append(row_vals)
            self._sheets[name] = rows_data

    # ------------------------------------------------------------------
    # Sheet access helpers
    # ------------------------------------------------------------------

    def _ensure_sheet(self, name: str) -> list[list[object]]:
        if name not in self._sheets:
            if name == "Review":
                self._sheets[name] = [list(REVIEW_COLUMNS)]
            else:
                self._sheets[name] = [list(MONTHLY_COLUMNS)]
        return self._sheets[name]

    def sheet_names(self) -> list[str]:
        return list(self._sheets.keys())

    # ------------------------------------------------------------------
    # Read existing keys
    # ------------------------------------------------------------------

    def read_keys(self) -> dict[tuple[str, int], tuple[str, int]]:
        result: dict[tuple[str, int], tuple[str, int]] = {}
        for sheet_name, rows in self._sheets.items():
            if sheet_name == "Review" or not rows:
                continue
            header = [str(h) if h is not None else "" for h in rows[0]]
            col_idx = {h: i for i, h in enumerate(header)}
            txnid_col = col_idx.get("Transaction ID")
            occ_col = col_idx.get("Occurrence")

            for row_num, row in enumerate(rows[1:], start=2):
                if txnid_col is None:
                    txn_id_val = row[5] if len(row) > 5 else None
                    occ_val: object = 0
                else:
                    txn_id_val = row[txnid_col] if txnid_col < len(row) else None
                    occ_val = row[occ_col] if occ_col is not None and occ_col < len(row) else 0

                if txn_id_val is not None and str(txn_id_val) != "":
                    txn_id_str = str(txn_id_val)
                    try:
                        occ_int = int(str(occ_val)) if occ_val is not None and occ_val != "" else 0
                    except (ValueError, TypeError):
                        occ_int = 0
                    result[(txn_id_str, occ_int)] = (sheet_name, row_num)
        return result

    # ------------------------------------------------------------------
    # Upsert
    # ------------------------------------------------------------------

    def upsert_row(
        self,
        txn: Transaction,
        sheet_name: str,
        existing_index: dict[tuple[str, int], tuple[str, int]],
    ) -> int:
        key = (txn.txn_id, txn.occurrence)
        if key in existing_index:
            return 0

        rows = self._ensure_sheet(sheet_name)
        now_str = datetime.datetime.now().isoformat(timespec="seconds")
        row_data: list[object] = [
            txn.date.isoformat() if txn.date else "",
            txn.description,
            str(txn.debit) if txn.debit is not None else "",
            str(txn.credit) if txn.credit is not None else "",
            txn.account,
            "",  # Category
            txn.txn_id,
            txn.occurrence,
            "",  # Document
            "",  # Archived Path
            "",  # Match Status
            "",  # Confidence
            now_str,
        ]
        rows.append(row_data)
        new_row_num = len(rows)  # 1-indexed; row 1 = header
        existing_index[key] = (sheet_name, new_row_num)
        return 1

    # ------------------------------------------------------------------
    # set_match
    # ------------------------------------------------------------------

    def set_match(
        self,
        key: tuple[str, int],
        document_link: str,
        archived_path: str,
        status: MatchStatus,
        confidence: float,
        existing_index: dict[tuple[str, int], tuple[str, int]],
    ) -> None:
        if key not in existing_index:
            return
        sheet_name, row_num = existing_index[key]
        rows = self._sheets.get(sheet_name)
        if rows is None or row_num < 2 or row_num > len(rows):
            return
        header = [str(h) for h in rows[0]]
        col_idx = {h: i for i, h in enumerate(header)}
        row = rows[row_num - 1]

        def _set(col_name: str, value: object) -> None:
            c = col_idx.get(col_name)
            if c is not None:
                while len(row) <= c:
                    row.append("")
                row[c] = value

        _set("Document", document_link)
        _set("Archived Path", archived_path)
        _set("Match Status", status.value)
        _set("Confidence", confidence)

    # ------------------------------------------------------------------
    # Review sheet
    # ------------------------------------------------------------------

    def add_to_review(self, items: list[str]) -> None:
        rows = self._ensure_sheet("Review")
        now_str = datetime.datetime.now().isoformat(timespec="seconds")
        for item in items:
            rows.append(["", "", item, now_str])

    # ------------------------------------------------------------------
    # Save
    # ------------------------------------------------------------------

    def save(self, target: Path) -> None:
        from odf import text as odf_text
        from odf.opendocument import OpenDocumentSpreadsheet
        from odf.table import Table, TableCell, TableRow

        doc = OpenDocumentSpreadsheet()

        for sheet_name, rows in self._sheets.items():
            sheet = Table(name=sheet_name)
            for row_data in rows:
                tr = TableRow()
                for cell_val in row_data:
                    tc = TableCell()
                    p = odf_text.P(text=str(cell_val) if cell_val is not None else "")
                    tc.addElement(p)
                    tr.addElement(tc)
                sheet.addElement(tr)
            doc.spreadsheet.addElement(sheet)

        doc.save(str(target))


# ---------------------------------------------------------------------------
# LedgerWorkbook — public class
# ---------------------------------------------------------------------------


class LedgerWorkbook:
    """Read, deduplicate, and write transaction data to a spreadsheet workbook."""

    def __init__(self, path: str | Path, fmt: Literal["xlsx", "ods"] = "xlsx") -> None:
        """Initialize without opening the file."""
        self._path = Path(path)
        self._fmt = fmt
        self._backend: _XlsxBackend | _OdsBackend | None = None
        # Keyed index: {(txn_id, occurrence): (sheet_name, row_number)}
        self._index: dict[tuple[str, int], tuple[str, int]] = {}

    # ------------------------------------------------------------------
    # open_or_create
    # ------------------------------------------------------------------

    def open_or_create(self) -> None:
        """Open existing workbook or create empty one. Never raises on missing file."""
        if self._fmt == "xlsx":
            backend: _XlsxBackend | _OdsBackend = _XlsxBackend()
        else:
            backend = _OdsBackend()
        backend.open_or_create(self._path)
        self._backend = backend
        self._index = backend.read_keys()

    # ------------------------------------------------------------------
    # existing_keys
    # ------------------------------------------------------------------

    def existing_keys(self) -> set[tuple[str, int]]:
        """Return set of (txn_id, occurrence) already in the workbook."""
        return set(self._index.keys())

    # ------------------------------------------------------------------
    # upsert
    # ------------------------------------------------------------------

    def upsert(self, transactions: list[Transaction]) -> tuple[int, int]:
        """Insert new transactions, skip existing ones by (txn_id, occurrence).

        Returns (rows_added, rows_updated). In v1 rows_updated is always 0.
        """
        assert self._backend is not None, "Call open_or_create() first"
        rows_added = 0
        for txn in transactions:
            sheet_name = f"{txn.date.year}-{txn.date.month:02d}"
            added = self._backend.upsert_row(txn, sheet_name, self._index)
            rows_added += added
        return rows_added, 0

    # ------------------------------------------------------------------
    # set_match
    # ------------------------------------------------------------------

    def set_match(
        self,
        key: tuple[str, int],
        document_link: str,
        archived_path: str,
        status: MatchStatus,
        confidence: float,
    ) -> None:
        """Update match-related cells for the row identified by (txn_id, occurrence)."""
        assert self._backend is not None, "Call open_or_create() first"
        self._backend.set_match(
            key, document_link, archived_path, status, confidence, self._index
        )

    # ------------------------------------------------------------------
    # add_to_review
    # ------------------------------------------------------------------

    def add_to_review(self, items: list[str]) -> None:
        """Add descriptive strings to the Review sheet."""
        assert self._backend is not None, "Call open_or_create() first"
        self._backend.add_to_review(items)

    # ------------------------------------------------------------------
    # save_atomic
    # ------------------------------------------------------------------

    def save_atomic(self) -> None:
        """Write to a temp file then rename over the target. Raises OutputWriteError on failure."""
        assert self._backend is not None, "Call open_or_create() first"
        path = self._path
        try:
            # Ensure parent directory exists — if it doesn't, we raise OutputWriteError
            if not path.parent.exists():
                raise OSError(f"Parent directory does not exist: {path.parent}")
            with tempfile.NamedTemporaryFile(
                dir=path.parent,
                delete=False,
                suffix=path.suffix,
            ) as tmp:
                tmp_name = tmp.name
            self._backend.save(Path(tmp_name))
            os.replace(tmp_name, path)
        except OSError as exc:
            # Clean up temp file if it was created
            try:
                os.unlink(tmp_name)
            except (OSError, NameError):
                pass
            raise OutputWriteError(f"Failed to write workbook to {path}: {exc}") from exc
