"""Tests for LedgerWorkbook (task 14)."""
from __future__ import annotations

import datetime
from decimal import Decimal
from pathlib import Path

import pytest

from docs_to_ledger.errors import OutputWriteError
from docs_to_ledger.ledger import LedgerWorkbook
from docs_to_ledger.model import MatchStatus, Transaction

FIXTURES = Path(__file__).parent / "fixtures"


def _txn(
    txn_id: str = "txn001",
    occurrence: int = 0,
    date: datetime.date = datetime.date(2026, 6, 15),
    description: str = "Test purchase",
    debit: Decimal | None = Decimal("50.00"),
    credit: Decimal | None = None,
    account: str = "Test Account",
    source_file: str = "test.csv",
) -> Transaction:
    return Transaction(
        date=date,
        description=description,
        debit=debit,
        credit=credit,
        account=account,
        source_file=source_file,
        txn_id=txn_id,
        occurrence=occurrence,
    )


# ---------------------------------------------------------------------------
# open_or_create
# ---------------------------------------------------------------------------


class TestOpenOrCreate:
    def test_non_existent_path_creates_new(self, tmp_path: Path) -> None:
        path = tmp_path / "new_ledger.xlsx"
        wb = LedgerWorkbook(path)
        wb.open_or_create()  # must not raise
        assert wb.existing_keys() == set()

    def test_existing_xlsx_opens_without_error(self, tmp_path: Path) -> None:
        import openpyxl  # type: ignore[import-untyped]

        path = tmp_path / "existing.xlsx"
        workbook = openpyxl.Workbook()
        workbook.save(str(path))

        wb = LedgerWorkbook(path)
        wb.open_or_create()  # must not raise

    def test_non_existent_ods_creates_new(self, tmp_path: Path) -> None:
        path = tmp_path / "new_ledger.ods"
        wb = LedgerWorkbook(path, fmt="ods")
        wb.open_or_create()  # must not raise
        assert wb.existing_keys() == set()


# ---------------------------------------------------------------------------
# upsert and existing_keys
# ---------------------------------------------------------------------------


class TestUpsertAndExistingKeys:
    def test_new_transaction_added(self, tmp_path: Path) -> None:
        path = tmp_path / "ledger.xlsx"
        wb = LedgerWorkbook(path)
        wb.open_or_create()
        txn = _txn()
        rows_added, rows_updated = wb.upsert([txn])
        assert rows_added == 1
        assert rows_updated == 0
        assert ("txn001", 0) in wb.existing_keys()

    def test_new_transaction_on_correct_sheet(self, tmp_path: Path) -> None:
        path = tmp_path / "ledger.xlsx"
        wb = LedgerWorkbook(path)
        wb.open_or_create()
        txn = _txn(date=datetime.date(2026, 6, 15))
        wb.upsert([txn])
        wb.save_atomic()

        import openpyxl  # type: ignore[import-untyped]

        workbook = openpyxl.load_workbook(str(path))
        assert "2026-06" in workbook.sheetnames

    def test_reimport_same_transaction_skipped(self, tmp_path: Path) -> None:
        path = tmp_path / "ledger.xlsx"
        wb = LedgerWorkbook(path)
        wb.open_or_create()
        txn = _txn()
        wb.upsert([txn])

        rows_added, rows_updated = wb.upsert([txn])
        assert rows_added == 0
        assert rows_updated == 0

    def test_same_txn_id_different_occurrence_both_added(self, tmp_path: Path) -> None:
        path = tmp_path / "ledger.xlsx"
        wb = LedgerWorkbook(path)
        wb.open_or_create()
        txn0 = _txn(txn_id="dup", occurrence=0)
        txn1 = _txn(txn_id="dup", occurrence=1)
        rows_added, _ = wb.upsert([txn0, txn1])
        assert rows_added == 2
        assert ("dup", 0) in wb.existing_keys()
        assert ("dup", 1) in wb.existing_keys()

    def test_two_transactions_different_accounts_same_date(self, tmp_path: Path) -> None:
        path = tmp_path / "ledger.xlsx"
        wb = LedgerWorkbook(path)
        wb.open_or_create()
        txn_a = _txn(txn_id="a", account="Account A")
        txn_b = _txn(txn_id="b", account="Account B")
        rows_added, _ = wb.upsert([txn_a, txn_b])
        assert rows_added == 2

    def test_transaction_creates_new_sheet_when_missing(self, tmp_path: Path) -> None:
        path = tmp_path / "ledger.xlsx"
        wb = LedgerWorkbook(path)
        wb.open_or_create()
        txn = _txn(date=datetime.date(2025, 12, 31))
        rows_added, _ = wb.upsert([txn])
        assert rows_added == 1
        wb.save_atomic()

        import openpyxl  # type: ignore[import-untyped]

        workbook = openpyxl.load_workbook(str(path))
        assert "2025-12" in workbook.sheetnames

    def test_transactions_on_different_months_go_to_different_sheets(
        self, tmp_path: Path
    ) -> None:
        path = tmp_path / "ledger.xlsx"
        wb = LedgerWorkbook(path)
        wb.open_or_create()
        txn_jun = _txn(txn_id="jun", date=datetime.date(2026, 6, 1))
        txn_jul = _txn(txn_id="jul", date=datetime.date(2026, 7, 1))
        rows_added, _ = wb.upsert([txn_jun, txn_jul])
        assert rows_added == 2
        wb.save_atomic()

        import openpyxl  # type: ignore[import-untyped]

        workbook = openpyxl.load_workbook(str(path))
        assert "2026-06" in workbook.sheetnames
        assert "2026-07" in workbook.sheetnames

    def test_existing_keys_empty_for_new_workbook(self, tmp_path: Path) -> None:
        path = tmp_path / "ledger.xlsx"
        wb = LedgerWorkbook(path)
        wb.open_or_create()
        assert wb.existing_keys() == set()

    def test_credit_transaction_added(self, tmp_path: Path) -> None:
        path = tmp_path / "ledger.xlsx"
        wb = LedgerWorkbook(path)
        wb.open_or_create()
        txn = _txn(txn_id="credit01", debit=None, credit=Decimal("100.00"))
        rows_added, _ = wb.upsert([txn])
        assert rows_added == 1


# ---------------------------------------------------------------------------
# Manual edit preservation
# ---------------------------------------------------------------------------


class TestManualEditPreservation:
    def test_category_preserved_on_re_upsert(self, tmp_path: Path) -> None:
        path = tmp_path / "ledger.xlsx"
        wb = LedgerWorkbook(path)
        wb.open_or_create()
        txn = _txn()
        wb.upsert([txn])
        wb.save_atomic()

        # Manually set Category value in the workbook
        import openpyxl  # type: ignore[import-untyped]

        workbook = openpyxl.load_workbook(str(path))
        ws = workbook["2026-06"]
        # Find Category column
        header_row = [cell.value for cell in ws[1]]
        cat_col = header_row.index("Category") + 1  # 1-indexed
        # Find data row (row 2)
        ws.cell(row=2, column=cat_col).value = "Groceries"
        workbook.save(str(path))

        # Re-open and re-upsert same transaction
        wb2 = LedgerWorkbook(path)
        wb2.open_or_create()
        rows_added, _ = wb2.upsert([txn])
        assert rows_added == 0  # skipped because already exists
        wb2.save_atomic()

        # Verify Category column is preserved
        workbook2 = openpyxl.load_workbook(str(path))
        ws2 = workbook2["2026-06"]
        header_row2 = [cell.value for cell in ws2[1]]
        cat_col2 = header_row2.index("Category") + 1
        assert ws2.cell(row=2, column=cat_col2).value == "Groceries"


# ---------------------------------------------------------------------------
# Review sheet
# ---------------------------------------------------------------------------


class TestReviewSheet:
    def test_add_to_review_creates_sheet(self, tmp_path: Path) -> None:
        path = tmp_path / "ledger.xlsx"
        wb = LedgerWorkbook(path)
        wb.open_or_create()
        wb.add_to_review(["some note"])
        wb.save_atomic()

        import openpyxl  # type: ignore[import-untyped]

        workbook = openpyxl.load_workbook(str(path))
        assert "Review" in workbook.sheetnames

    def test_add_to_review_content(self, tmp_path: Path) -> None:
        path = tmp_path / "ledger.xlsx"
        wb = LedgerWorkbook(path)
        wb.open_or_create()
        wb.add_to_review(["note one", "note two"])
        wb.save_atomic()

        import openpyxl  # type: ignore[import-untyped]

        workbook = openpyxl.load_workbook(str(path))
        ws = workbook["Review"]
        values = [ws.cell(row=r, column=3).value for r in range(2, ws.max_row + 1)]
        assert "note one" in values
        assert "note two" in values

    def test_add_to_review_appends_on_multiple_calls(self, tmp_path: Path) -> None:
        path = tmp_path / "ledger.xlsx"
        wb = LedgerWorkbook(path)
        wb.open_or_create()
        wb.add_to_review(["first"])
        wb.add_to_review(["second"])
        wb.save_atomic()

        import openpyxl  # type: ignore[import-untyped]

        workbook = openpyxl.load_workbook(str(path))
        ws = workbook["Review"]
        # Should have header + 2 data rows
        data_rows = ws.max_row - 1
        assert data_rows == 2


# ---------------------------------------------------------------------------
# set_match
# ---------------------------------------------------------------------------


class TestSetMatch:
    def test_set_match_updates_cells(self, tmp_path: Path) -> None:
        path = tmp_path / "ledger.xlsx"
        wb = LedgerWorkbook(path)
        wb.open_or_create()
        txn = _txn()
        wb.upsert([txn])
        wb.set_match(
            key=("txn001", 0),
            document_link="https://example.com/doc.pdf",
            archived_path="/archive/2026/06/doc.pdf",
            status=MatchStatus.matched,
            confidence=0.95,
        )
        wb.save_atomic()

        import openpyxl  # type: ignore[import-untyped]

        workbook = openpyxl.load_workbook(str(path))
        ws = workbook["2026-06"]
        header = [cell.value for cell in ws[1]]
        doc_col = header.index("Document") + 1
        arch_col = header.index("Archived Path") + 1
        status_col = header.index("Match Status") + 1
        conf_col = header.index("Confidence") + 1

        assert ws.cell(row=2, column=doc_col).value == "https://example.com/doc.pdf"
        assert ws.cell(row=2, column=arch_col).value == "/archive/2026/06/doc.pdf"
        assert ws.cell(row=2, column=status_col).value == "matched"
        assert ws.cell(row=2, column=conf_col).value == pytest.approx(0.95)

    def test_set_match_unknown_key_no_error(self, tmp_path: Path) -> None:
        path = tmp_path / "ledger.xlsx"
        wb = LedgerWorkbook(path)
        wb.open_or_create()
        # Should not raise even if key is not found
        wb.set_match(
            key=("nonexistent", 0),
            document_link="https://example.com/doc.pdf",
            archived_path="/archive/doc.pdf",
            status=MatchStatus.unmatched,
            confidence=0.0,
        )


# ---------------------------------------------------------------------------
# Atomic write
# ---------------------------------------------------------------------------


class TestAtomicWrite:
    def test_save_atomic_creates_file(self, tmp_path: Path) -> None:
        path = tmp_path / "ledger.xlsx"
        wb = LedgerWorkbook(path)
        wb.open_or_create()
        wb.upsert([_txn()])
        wb.save_atomic()
        assert path.exists()

    def test_save_atomic_unwritable_raises(self, tmp_path: Path) -> None:
        # Use a directory as path to trigger write failure
        path = tmp_path / "unwritable_dir" / "ledger.xlsx"
        # Parent directory does not exist → write will fail
        wb = LedgerWorkbook(path)
        wb.open_or_create()
        wb.upsert([_txn()])
        with pytest.raises(OutputWriteError):
            wb.save_atomic()


# ---------------------------------------------------------------------------
# Format round-trip
# ---------------------------------------------------------------------------


class TestFormatRoundTrip:
    def test_xlsx_round_trip(self, tmp_path: Path) -> None:
        path = tmp_path / "ledger.xlsx"
        wb = LedgerWorkbook(path)
        wb.open_or_create()
        txn = _txn(txn_id="rt001")
        wb.upsert([txn])
        wb.save_atomic()

        wb2 = LedgerWorkbook(path)
        wb2.open_or_create()
        assert ("rt001", 0) in wb2.existing_keys()

    def test_ods_round_trip(self, tmp_path: Path) -> None:
        path = tmp_path / "ledger.ods"
        wb = LedgerWorkbook(path, fmt="ods")
        wb.open_or_create()
        txn = _txn(txn_id="rt_ods001")
        wb.upsert([txn])
        wb.save_atomic()

        wb2 = LedgerWorkbook(path, fmt="ods")
        wb2.open_or_create()
        assert ("rt_ods001", 0) in wb2.existing_keys()

    def test_xlsx_round_trip_preserves_values(self, tmp_path: Path) -> None:
        path = tmp_path / "ledger.xlsx"
        wb = LedgerWorkbook(path)
        wb.open_or_create()
        txn = _txn(
            txn_id="rt002",
            description="Round trip test",
            debit=Decimal("123.45"),
            account="Test Account",
        )
        wb.upsert([txn])
        wb.save_atomic()

        import openpyxl  # type: ignore[import-untyped]

        workbook = openpyxl.load_workbook(str(path))
        ws = workbook["2026-06"]
        header = [cell.value for cell in ws[1]]
        desc_col = header.index("Description") + 1
        assert ws.cell(row=2, column=desc_col).value == "Round trip test"


# ---------------------------------------------------------------------------
# Dedupe with existing fixture (old 7-column format)
# ---------------------------------------------------------------------------


class TestFixtureDedupe:
    def test_existing_fixture_keys_loaded(self) -> None:
        fixture = FIXTURES / "sample_workbook.xlsx"
        wb = LedgerWorkbook(fixture)
        wb.open_or_create()
        keys = wb.existing_keys()
        assert ("abc123", 0) in keys

    def test_upsert_existing_fixture_row_skipped(self, tmp_path: Path) -> None:
        import shutil

        fixture = FIXTURES / "sample_workbook.xlsx"
        path = tmp_path / "workbook.xlsx"
        shutil.copy(fixture, path)

        wb = LedgerWorkbook(path)
        wb.open_or_create()

        txn = Transaction(
            date=datetime.date(2026, 6, 15),
            description="EXISTING ROW",
            debit=Decimal("149.90"),
            credit=None,
            account="SG Compte Courant",
            source_file="test.csv",
            txn_id="abc123",
            occurrence=0,
        )
        rows_added, _ = wb.upsert([txn])
        assert rows_added == 0

    def test_old_format_workbook_opens_gracefully(self) -> None:
        fixture = FIXTURES / "sample_workbook.xlsx"
        wb = LedgerWorkbook(fixture)
        wb.open_or_create()  # must not raise
        # Should have loaded at least 1 key
        assert len(wb.existing_keys()) >= 1

    def test_upsert_into_old_format_adds_new_transaction(self, tmp_path: Path) -> None:
        import shutil

        fixture = FIXTURES / "sample_workbook.xlsx"
        path = tmp_path / "workbook.xlsx"
        shutil.copy(fixture, path)

        wb = LedgerWorkbook(path)
        wb.open_or_create()

        new_txn = _txn(txn_id="new999", date=datetime.date(2026, 6, 20))
        rows_added, _ = wb.upsert([new_txn])
        assert rows_added == 1
        wb.save_atomic()

        # Verify the new transaction can be reloaded
        wb2 = LedgerWorkbook(path)
        wb2.open_or_create()
        assert ("new999", 0) in wb2.existing_keys()


# ---------------------------------------------------------------------------
# Edge cases
# ---------------------------------------------------------------------------


class TestEdgeCases:
    def test_upsert_empty_list(self, tmp_path: Path) -> None:
        path = tmp_path / "ledger.xlsx"
        wb = LedgerWorkbook(path)
        wb.open_or_create()
        rows_added, rows_updated = wb.upsert([])
        assert rows_added == 0
        assert rows_updated == 0

    def test_upsert_returns_correct_count_mixed(self, tmp_path: Path) -> None:
        path = tmp_path / "ledger.xlsx"
        wb = LedgerWorkbook(path)
        wb.open_or_create()
        txn_a = _txn(txn_id="mix_a")
        txn_b = _txn(txn_id="mix_b")
        wb.upsert([txn_a])  # first insert

        # Second upsert: txn_a already exists, txn_b is new
        rows_added, rows_updated = wb.upsert([txn_a, txn_b])
        assert rows_added == 1
        assert rows_updated == 0

    def test_path_as_string(self, tmp_path: Path) -> None:
        path = str(tmp_path / "ledger.xlsx")
        wb = LedgerWorkbook(path)
        wb.open_or_create()  # must not raise

    def test_multiple_save_atomic_calls(self, tmp_path: Path) -> None:
        path = tmp_path / "ledger.xlsx"
        wb = LedgerWorkbook(path)
        wb.open_or_create()
        wb.upsert([_txn(txn_id="s1")])
        wb.save_atomic()
        wb.upsert([_txn(txn_id="s2")])
        wb.save_atomic()

        wb2 = LedgerWorkbook(path)
        wb2.open_or_create()
        assert ("s1", 0) in wb2.existing_keys()
        assert ("s2", 0) in wb2.existing_keys()
