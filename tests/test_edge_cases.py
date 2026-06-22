"""Edge-case and error tests — one per spec bullet.

Edge cases covered:
  EC-1: Statement row that fails to parse → skipped, reported in parse_failures
  EC-2: File not attributable to any account → reported as unrouted
  EC-3: Legitimately identical transactions → all kept via occurrence index
  EC-4: Document matches multiple transactions → ambiguous → Review, not linked
  EC-5: Multiple documents for one transaction → best linked, rest to Review
  EC-6: OCR returns no usable amount/date → document goes to Review
  EC-7: Manual edits preserved across re-run (dedupe on txn_id, not full row)
  EC-8: Same document already archived → not re-copied on re-run
  EC-9: Credit transaction matched by absolute amount

Errors covered:
  E-1: Invalid YAML config → ConfigError before any file touched
  E-2: Tesseract not installed + image/OCR doc present → OcrUnavailableError propagates
  E-3: Text-only run proceeds even when Tesseract is unavailable
  E-4: Output directory not writable → OutputWriteError propagates
"""
from __future__ import annotations

import datetime
from decimal import Decimal
from pathlib import Path
from unittest.mock import patch

import openpyxl
import pytest

from docs_to_ledger.errors import ConfigError, OcrUnavailableError, OutputWriteError
from docs_to_ledger.model import DocumentFeature
from docs_to_ledger.runner import RunArgs, run

# ---------------------------------------------------------------------------
# Shared helpers
# ---------------------------------------------------------------------------

_CONFIG = """\
sources:
  - account: "Test Account"
    statement_type: csv
    path_pattern: "statements/*.csv"
    locale:
      date_format: "dd/mm/yyyy"
      decimal_separator: comma
      thousands_separator: space
    column_map:
      date: "Date"
      description: "Libelle"
      debit: "Debit"
      credit: "Credit"
matching:
  date_window_days: 3
  amount_tolerance: 0
  fuzzy_vendor: true
archive:
  root: "archive"
  layout: "year/month"
"""


def _build_tree(tmp_path: Path, csv_content: str) -> dict[str, Path]:
    config_path = tmp_path / "config.yaml"
    config_path.write_text(_CONFIG, encoding="utf-8")
    stmts = tmp_path / "input" / "statements"
    stmts.mkdir(parents=True)
    (stmts / "bank.csv").write_text(csv_content, encoding="utf-8")
    return {
        "config_path": config_path,
        "input_path": tmp_path / "input",
        "output_dir": tmp_path / "output",
    }


def _args(tree: dict[str, Path], *, dry_run: bool = False) -> RunArgs:
    return RunArgs(
        input_path=tree["input_path"],
        config_path=tree["config_path"],
        output_dir=tree["output_dir"],
        dry_run=dry_run,
    )


# ---------------------------------------------------------------------------
# EC-1: Statement row with bad date is skipped and reported
# ---------------------------------------------------------------------------


def test_bad_date_row_skipped_and_reported(tmp_path: Path) -> None:
    """Spec EC-1: row that fails to parse → skipped, logged with file+line."""
    csv = (
        "Date;Libelle;Debit;Credit\n"
        "NOT_A_DATE;SUPER U;149,90;\n"  # bad date → parse failure
        "15/06/2026;VIREMENT;;2500,00\n"  # valid row
    )
    tree = _build_tree(tmp_path, csv)
    with patch("docs_to_ledger.runner.extract_features", return_value=None):
        report = run(_args(tree))

    assert report.parse_failures == 1
    assert report.rows_added == 1  # only the valid row was added


def test_bad_amount_row_skipped_and_reported(tmp_path: Path) -> None:
    """Spec EC-1: row with unparseable amount → skipped, logged."""
    csv = (
        "Date;Libelle;Debit;Credit\n"
        "15/06/2026;SUPER U;INVALID_AMT;\n"  # bad amount → parse failure
        "14/06/2026;VIREMENT;;2500,00\n"
    )
    tree = _build_tree(tmp_path, csv)
    with patch("docs_to_ledger.runner.extract_features", return_value=None):
        report = run(_args(tree))

    assert report.parse_failures == 1
    assert report.rows_added == 1


# ---------------------------------------------------------------------------
# EC-2: Unrouted file reported and skipped
# ---------------------------------------------------------------------------


def test_unrouted_file_counted_in_report(tmp_path: Path) -> None:
    """Spec EC-2: file with no routing rule → reported as 'unrouted'."""
    tree = _build_tree(tmp_path, "Date;Libelle;Debit;Credit\n15/06/2026;X;10,00;\n")
    # Place a CSV that does not match "statements/*.csv"
    other = tree["input_path"] / "other_dir" / "bank.csv"
    other.parent.mkdir()
    other.write_text("Date;Libelle\n15/06/2026;something\n", encoding="utf-8")

    with patch("docs_to_ledger.runner.extract_features", return_value=None):
        report = run(_args(tree))

    assert report.unrouted >= 1


# ---------------------------------------------------------------------------
# EC-3: Identical transactions kept via occurrence index
# ---------------------------------------------------------------------------


def test_identical_transactions_kept_with_occurrence_index(tmp_path: Path) -> None:
    """Spec EC-3: two rows identical in date/amount/desc → both kept (occurrence 0 and 1)."""
    csv = (
        "Date;Libelle;Debit;Credit\n"
        "15/06/2026;SUPER U;149,90;\n"
        "15/06/2026;SUPER U;149,90;\n"  # exact duplicate
    )
    tree = _build_tree(tmp_path, csv)
    with patch("docs_to_ledger.runner.extract_features", return_value=None):
        report = run(_args(tree))

    assert report.rows_added == 2


def test_identical_transactions_not_duplicated_on_rerun(tmp_path: Path) -> None:
    """Spec EC-3: re-importing same statement does not create duplicates."""
    csv = (
        "Date;Libelle;Debit;Credit\n"
        "15/06/2026;SUPER U;149,90;\n"
        "15/06/2026;SUPER U;149,90;\n"
    )
    tree = _build_tree(tmp_path, csv)
    with patch("docs_to_ledger.runner.extract_features", return_value=None):
        run(_args(tree))
        report2 = run(_args(tree))

    assert report2.rows_added == 0  # both occurrence 0 and 1 already present


# ---------------------------------------------------------------------------
# EC-4: Document matches multiple transactions → ambiguous → Review, not linked
# ---------------------------------------------------------------------------


def test_doc_matches_multiple_txns_is_ambiguous(tmp_path: Path) -> None:
    """Spec EC-4: one doc, two txns same amount+date, no vendor → ambiguous, not linked."""
    csv = (
        "Date;Libelle;Debit;Credit\n"
        "15/06/2026;SUPER U;149,90;\n"
        "15/06/2026;CARREFOUR;149,90;\n"  # same amount and date, different vendor
    )
    tree = _build_tree(tmp_path, csv)
    docs_dir = tree["input_path"] / "docs"
    docs_dir.mkdir()
    (docs_dir / "receipt.pdf").write_bytes(b"%PDF-1.4 receipt")

    ambiguous_doc = DocumentFeature(
        source_path=str(docs_dir / "receipt.pdf"),
        amount=Decimal("149.90"),
        date=datetime.date(2026, 6, 15),
        vendor=None,  # no vendor → cannot disambiguate
        needs_ocr=False,
    )
    with patch("docs_to_ledger.runner.extract_features", return_value=ambiguous_doc):
        report = run(_args(tree))

    assert report.docs_matched == 0  # ambiguous doc is not linked
    assert report.review_items > 0   # doc and/or txns appear on Review


# ---------------------------------------------------------------------------
# EC-5: Multiple documents for one transaction → best linked, rest to Review
# ---------------------------------------------------------------------------


def test_multiple_docs_for_one_txn_best_wins(tmp_path: Path) -> None:
    """Spec EC-5: two docs match one txn → higher-confidence winner linked, other to Review."""
    csv = "Date;Libelle;Debit;Credit\n15/06/2026;SUPER U;149,90;\n"
    tree = _build_tree(tmp_path, csv)
    docs_dir = tree["input_path"] / "docs"
    docs_dir.mkdir()
    (docs_dir / "receipt_a.pdf").write_bytes(b"%PDF-1.4 a")
    (docs_dir / "receipt_b.pdf").write_bytes(b"%PDF-1.4 b")

    doc_a = DocumentFeature(
        source_path=str(docs_dir / "receipt_a.pdf"),
        amount=Decimal("149.90"),
        date=datetime.date(2026, 6, 15),
        vendor="SUPER U",
        needs_ocr=False,
    )
    doc_b = DocumentFeature(
        source_path=str(docs_dir / "receipt_b.pdf"),
        amount=Decimal("149.90"),
        date=datetime.date(2026, 6, 15),
        vendor="SUPER U",
        needs_ocr=False,
    )

    with patch("docs_to_ledger.runner.extract_features", side_effect=[doc_a, doc_b]):
        report = run(_args(tree))

    assert report.docs_matched == 1  # exactly one doc linked to the transaction
    assert report.review_items >= 1  # the losing doc ends up on Review


# ---------------------------------------------------------------------------
# EC-6: OCR returns no usable amount/date → document goes to Review
# ---------------------------------------------------------------------------


def test_ocr_no_amount_date_doc_goes_to_review(tmp_path: Path) -> None:
    """Spec EC-6: doc with no parseable amount/date → Review as unmatchable."""
    csv = "Date;Libelle;Debit;Credit\n15/06/2026;SUPER U;149,90;\n"
    tree = _build_tree(tmp_path, csv)
    docs_dir = tree["input_path"] / "docs"
    docs_dir.mkdir()
    (docs_dir / "unreadable.pdf").write_bytes(b"%PDF-1.4 unreadable")

    sentinel = DocumentFeature(
        source_path=str(docs_dir / "unreadable.pdf"),
        amount=Decimal("0"),
        date=None,
        vendor=None,
        needs_ocr=True,  # signals OCR/extraction failure
    )
    with patch("docs_to_ledger.runner.extract_features", return_value=sentinel):
        report = run(_args(tree))

    assert report.docs_matched == 0
    assert report.review_items >= 2  # unmatched txn + unmatchable doc


# ---------------------------------------------------------------------------
# EC-7: Manual edits preserved across re-run
# ---------------------------------------------------------------------------


def test_manual_edit_preserved_across_rerun(tmp_path: Path) -> None:
    """Spec EC-7: manual Category edit is preserved because dedupe skips existing rows."""
    csv = "Date;Libelle;Debit;Credit\n15/06/2026;SUPER U;149,90;\n"
    tree = _build_tree(tmp_path, csv)

    with patch("docs_to_ledger.runner.extract_features", return_value=None):
        run(_args(tree))

    # Manually set Category in the workbook
    wb_path = tree["output_dir"] / "Test Account.xlsx"
    wb = openpyxl.load_workbook(str(wb_path))
    ws = wb["2026-06"]
    header = [ws.cell(1, c).value for c in range(1, ws.max_column + 1)]
    cat_col = header.index("Category") + 1  # 1-based
    ws.cell(2, cat_col).value = "Groceries"
    wb.save(str(wb_path))

    with patch("docs_to_ledger.runner.extract_features", return_value=None):
        report2 = run(_args(tree))

    assert report2.rows_added == 0  # row was already there — deduped

    wb2 = openpyxl.load_workbook(str(wb_path))
    ws2 = wb2["2026-06"]
    assert ws2.cell(2, cat_col).value == "Groceries"  # edit survived


# ---------------------------------------------------------------------------
# EC-8: Same document already archived → not re-copied on re-run
# ---------------------------------------------------------------------------


def test_already_archived_doc_not_recopied(tmp_path: Path) -> None:
    """Spec EC-8: re-run on a matched doc returns existing archive path, no extra copy."""
    csv = "Date;Libelle;Debit;Credit\n15/06/2026;SUPER U;149,90;\n"
    tree = _build_tree(tmp_path, csv)
    docs_dir = tree["input_path"] / "docs"
    docs_dir.mkdir()
    receipt = docs_dir / "receipt.pdf"
    receipt.write_bytes(b"%PDF-1.4 receipt content")

    doc = DocumentFeature(
        source_path=str(receipt),
        amount=Decimal("149.90"),
        date=datetime.date(2026, 6, 15),
        vendor="SUPER U",
        needs_ocr=False,
    )

    with patch("docs_to_ledger.runner.extract_features", return_value=doc):
        run(_args(tree))
        run(_args(tree))  # second run

    archive_pdfs = list((tree["output_dir"] / "archive").rglob("*.pdf"))
    assert len(archive_pdfs) == 1  # no duplicate copies


# ---------------------------------------------------------------------------
# EC-9: Credit transaction matched by absolute amount
# ---------------------------------------------------------------------------


def test_credit_txn_matched_by_absolute_amount(tmp_path: Path) -> None:
    """Spec EC-9: credit transaction (money in) is matched via absolute amount comparison."""
    csv = "Date;Libelle;Debit;Credit\n14/06/2026;VIREMENT SALAIRE;;2500,00\n"
    tree = _build_tree(tmp_path, csv)
    # Candidate doc must be discoverable inside input_path
    docs_dir = tree["input_path"] / "docs"
    docs_dir.mkdir()
    payslip = docs_dir / "payslip.pdf"
    payslip.write_bytes(b"%PDF-1.4 payslip")

    doc = DocumentFeature(
        source_path=str(payslip),
        amount=Decimal("2500.00"),
        date=datetime.date(2026, 6, 14),
        vendor="VIREMENT SALAIRE",
        needs_ocr=False,
    )
    with patch("docs_to_ledger.runner.extract_features", return_value=doc):
        report = run(_args(tree))

    assert report.docs_matched == 1


# ---------------------------------------------------------------------------
# E-1: Invalid YAML config → ConfigError (fail fast, no files touched)
# ---------------------------------------------------------------------------


def test_invalid_yaml_raises_config_error(tmp_path: Path) -> None:
    """Spec E-1: invalid YAML in config → ConfigError before any file is touched."""
    config_path = tmp_path / "bad.yaml"
    config_path.write_text("sources: [\n  unclosed: bracket\n", encoding="utf-8")

    with pytest.raises(ConfigError):
        run(RunArgs(input_path=tmp_path, config_path=config_path))


def test_missing_sources_raises_config_error(tmp_path: Path) -> None:
    """Spec E-1: config without 'sources' → ConfigError."""
    config_path = tmp_path / "no_sources.yaml"
    config_path.write_text("matching:\n  date_window_days: 3\n", encoding="utf-8")

    with pytest.raises(ConfigError):
        run(RunArgs(input_path=tmp_path, config_path=config_path))


# ---------------------------------------------------------------------------
# E-2: Tesseract not installed + image present → OcrUnavailableError propagates
# ---------------------------------------------------------------------------


def test_ocr_unavailable_propagates_when_image_present(tmp_path: Path) -> None:
    """Spec E-2: Tesseract missing + image doc → OcrUnavailableError from runner."""
    csv = "Date;Libelle;Debit;Credit\n15/06/2026;SUPER U;149,90;\n"
    tree = _build_tree(tmp_path, csv)
    img = tree["input_path"] / "receipt.png"
    img.write_bytes(b"\x89PNG\r\n\x1a\n")  # minimal PNG header

    with (
        patch(
            "docs_to_ledger.documents.ocr.ensure_available",
            side_effect=OcrUnavailableError("Tesseract not found"),
        ),
        pytest.raises(OcrUnavailableError),
    ):
        run(_args(tree))


# ---------------------------------------------------------------------------
# E-3: Text-only run proceeds even when Tesseract is unavailable
# ---------------------------------------------------------------------------


def test_text_only_run_proceeds_without_tesseract(tmp_path: Path) -> None:
    """Spec E-3: when no OCR path is taken, Tesseract unavailability is irrelevant."""
    csv = "Date;Libelle;Debit;Credit\n15/06/2026;SUPER U;149,90;\n"
    tree = _build_tree(tmp_path, csv)
    # Candidate doc must be inside input_path to be discovered
    docs_dir = tree["input_path"] / "docs"
    docs_dir.mkdir()
    invoice = docs_dir / "invoice.pdf"
    invoice.write_bytes(b"%PDF-1.4 invoice")

    text_feature = DocumentFeature(
        source_path=str(invoice),
        amount=Decimal("149.90"),
        date=datetime.date(2026, 6, 15),
        vendor="SUPER U",
        needs_ocr=False,  # extracted from text PDF — no OCR call made
    )
    with (
        patch("docs_to_ledger.runner.extract_features", return_value=text_feature),
        patch(
            "docs_to_ledger.ocr.ensure_available",
            side_effect=OcrUnavailableError("Tesseract not found"),
        ),
    ):
        # Should NOT raise — OCR path was never taken
        report = run(_args(tree))

    assert report.docs_matched == 1


# ---------------------------------------------------------------------------
# E-4: Output directory not writable → OutputWriteError propagates
# ---------------------------------------------------------------------------


def test_output_not_writable_raises_output_write_error(tmp_path: Path) -> None:
    """Spec E-4: unwritable output → OutputWriteError with no partial corruption."""
    csv = "Date;Libelle;Debit;Credit\n15/06/2026;SUPER U;149,90;\n"
    tree = _build_tree(tmp_path, csv)

    with (
        patch("docs_to_ledger.runner.extract_features", return_value=None),
        patch(
            "docs_to_ledger.ledger.LedgerWorkbook.save_atomic",
            side_effect=OutputWriteError("disk full"),
        ),
        pytest.raises(OutputWriteError),
    ):
        run(_args(tree))
