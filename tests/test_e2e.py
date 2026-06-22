"""Happy-path end-to-end tests.

Covers:
  E2E-1: Multi-account input tree → correct workbooks, archive, report counters.
  E2E-2: Second run over the same tree is fully idempotent (no new rows added,
          no extra archive copies).
"""
from __future__ import annotations

import datetime
from decimal import Decimal
from pathlib import Path
from unittest.mock import patch

import openpyxl
import pytest

from docs_to_ledger.model import DocumentFeature
from docs_to_ledger.runner import RunArgs, run

# ---------------------------------------------------------------------------
# Fixture: two-account input tree
# ---------------------------------------------------------------------------

_CONFIG = """\
sources:
  - account: "Account A"
    statement_type: csv
    path_pattern: "acct_a/statements/*.csv"
    locale:
      date_format: "dd/mm/yyyy"
      decimal_separator: comma
      thousands_separator: space
    column_map:
      date: "Date"
      description: "Libelle"
      debit: "Debit"
      credit: "Credit"
  - account: "Account B"
    statement_type: csv
    path_pattern: "acct_b/statements/*.csv"
    locale:
      date_format: "dd/mm/yyyy"
      decimal_separator: comma
      thousands_separator: space
    column_map:
      date: "Date"
      description: "Libelle"
      debit: "Debit"
      credit: "Credit"
matching:
  date_window_days: 3
  amount_tolerance: 0
  fuzzy_vendor: true
archive:
  root: "archive"
  layout: "year/month"
"""

_CSV_A = (
    "Date;Libelle;Debit;Credit\n"
    "15/06/2026;SUPER U;149,90;\n"
    "14/06/2026;VIREMENT SALAIRE;;2500,00\n"
)

_CSV_B = (
    "Date;Libelle;Debit;Credit\n"
    "20/06/2026;LOYER;800,00;\n"
)


@pytest.fixture()
def e2e_tree(tmp_path: Path) -> dict[str, Path]:
    """Two-account input tree: two statement CSVs and three receipt PDFs."""
    (tmp_path / "config.yaml").write_text(_CONFIG, encoding="utf-8")

    stmts_a = tmp_path / "input" / "acct_a" / "statements"
    stmts_a.mkdir(parents=True)
    (stmts_a / "june.csv").write_text(_CSV_A, encoding="utf-8")

    stmts_b = tmp_path / "input" / "acct_b" / "statements"
    stmts_b.mkdir(parents=True)
    (stmts_b / "june.csv").write_text(_CSV_B, encoding="utf-8")

    docs_dir = tmp_path / "input" / "docs"
    docs_dir.mkdir()
    (docs_dir / "invoice_loyer.pdf").write_bytes(b"%PDF-1.4 loyer")
    (docs_dir / "receipt_a.pdf").write_bytes(b"%PDF-1.4 receipt_a")
    (docs_dir / "receipt_b.pdf").write_bytes(b"%PDF-1.4 receipt_b")

    return {
        "config_path": tmp_path / "config.yaml",
        "input_path": tmp_path / "input",
        "output_dir": tmp_path / "output",
        "docs_dir": docs_dir,
    }


def _args(tree: dict[str, Path]) -> RunArgs:
    return RunArgs(
        input_path=tree["input_path"],
        config_path=tree["config_path"],
        output_dir=tree["output_dir"],
    )


def _mock_ef(docs_dir: Path):
    """Return a mock side_effect function that maps doc paths to DocumentFeatures."""
    features = {
        str(docs_dir / "invoice_loyer.pdf"): DocumentFeature(
            source_path=str(docs_dir / "invoice_loyer.pdf"),
            amount=Decimal("800.00"),
            date=datetime.date(2026, 6, 20),
            vendor="LOYER",
            needs_ocr=False,
        ),
        str(docs_dir / "receipt_a.pdf"): DocumentFeature(
            source_path=str(docs_dir / "receipt_a.pdf"),
            amount=Decimal("149.90"),
            date=datetime.date(2026, 6, 15),
            vendor="SUPER U",
            needs_ocr=False,
        ),
        str(docs_dir / "receipt_b.pdf"): DocumentFeature(
            source_path=str(docs_dir / "receipt_b.pdf"),
            amount=Decimal("2500.00"),
            date=datetime.date(2026, 6, 14),
            vendor="VIREMENT SALAIRE",
            needs_ocr=False,
        ),
    }

    def _ef(path: object) -> DocumentFeature | None:
        return features.get(str(path))

    return _ef


# ---------------------------------------------------------------------------
# E2E-1: Happy-path multi-account run
# ---------------------------------------------------------------------------


def test_e2e_multi_account_happy_path(e2e_tree: dict[str, Path]) -> None:
    """Both accounts produce correct workbooks, archive, and report counters."""
    ef = _mock_ef(e2e_tree["docs_dir"])
    with patch("docs_to_ledger.runner.extract_features", side_effect=ef):
        report = run(_args(e2e_tree))

    out = e2e_tree["output_dir"]

    # Both workbooks exist
    assert (out / "Account A.xlsx").exists()
    assert (out / "Account B.xlsx").exists()

    # Correct rows in Account A
    wb_a = openpyxl.load_workbook(str(out / "Account A.xlsx"))
    assert "2026-06" in wb_a.sheetnames
    ws_a = wb_a["2026-06"]
    assert ws_a.max_row == 3  # header + 2 transactions

    # Correct rows in Account B
    wb_b = openpyxl.load_workbook(str(out / "Account B.xlsx"))
    assert "2026-06" in wb_b.sheetnames
    ws_b = wb_b["2026-06"]
    assert ws_b.max_row == 2  # header + 1 transaction

    # Archive created under year/month layout
    archive = out / "archive" / "2026"
    assert archive.exists()
    all_archived = list(archive.rglob("*.pdf"))
    assert len(all_archived) == 3  # one copy per matched doc

    # Report counters
    assert report.rows_added == 3       # 2 for A, 1 for B
    assert report.rows_updated == 0
    assert report.docs_matched == 3     # receipt_a + receipt_b for A; invoice_loyer for B
    assert report.docs_copied == 3
    assert report.parse_failures == 0
    assert report.unrouted == 0


def test_e2e_matched_rows_have_match_status(e2e_tree: dict[str, Path]) -> None:
    """Matched transaction rows have Match Status = 'matched'."""
    ef = _mock_ef(e2e_tree["docs_dir"])
    with patch("docs_to_ledger.runner.extract_features", side_effect=ef):
        run(_args(e2e_tree))

    out = e2e_tree["output_dir"]
    wb_a = openpyxl.load_workbook(str(out / "Account A.xlsx"))
    ws_a = wb_a["2026-06"]
    header = [ws_a.cell(1, c).value for c in range(1, ws_a.max_column + 1)]
    status_col = header.index("Match Status") + 1

    statuses = {ws_a.cell(r, status_col).value for r in range(2, ws_a.max_row + 1)}
    assert "matched" in statuses


def test_e2e_archive_layout_is_year_month(e2e_tree: dict[str, Path]) -> None:
    """Archive follows year/month directory layout."""
    ef = _mock_ef(e2e_tree["docs_dir"])
    with patch("docs_to_ledger.runner.extract_features", side_effect=ef):
        run(_args(e2e_tree))

    archive_root = e2e_tree["output_dir"] / "archive"
    assert (archive_root / "2026" / "06").is_dir()


# ---------------------------------------------------------------------------
# E2E-2: Second run is fully idempotent
# ---------------------------------------------------------------------------


def test_e2e_second_run_adds_no_rows(e2e_tree: dict[str, Path]) -> None:
    """Spec: re-running adds 0 rows (Transaction ID dedupe prevents duplication)."""
    mock_ef = _mock_ef(e2e_tree["docs_dir"])
    with patch("docs_to_ledger.runner.extract_features", side_effect=mock_ef):
        run(_args(e2e_tree))

    # Reset side_effect for second run
    mock_ef2 = _mock_ef(e2e_tree["docs_dir"])
    with patch("docs_to_ledger.runner.extract_features", side_effect=mock_ef2):
        report2 = run(_args(e2e_tree))

    assert report2.rows_added == 0


def test_e2e_second_run_no_extra_archive_files(e2e_tree: dict[str, Path]) -> None:
    """Spec: re-run does not create additional archive copies (content check in archive module)."""
    mock_ef = _mock_ef(e2e_tree["docs_dir"])
    with patch("docs_to_ledger.runner.extract_features", side_effect=mock_ef):
        run(_args(e2e_tree))

    archive = e2e_tree["output_dir"] / "archive"
    count_after_run1 = len(list(archive.rglob("*.pdf")))

    mock_ef2 = _mock_ef(e2e_tree["docs_dir"])
    with patch("docs_to_ledger.runner.extract_features", side_effect=mock_ef2):
        run(_args(e2e_tree))

    count_after_run2 = len(list(archive.rglob("*.pdf")))
    assert count_after_run2 == count_after_run1
